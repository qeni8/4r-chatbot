"""xlsx → atik_kodlari tablosu (Brief Bölüm 5).

Kullanım:
    python scripts/ingest_atik_kodlari.py data/raw/GUNCEL_ATIK_KODLARI_4R.xlsx

Excel düzeni (doğrulandı, 974 satır / 842 adet 6-haneli kod):
    A: ATIK KODU  B: TANIM  C: AÇIKLAMA(A/M, yok sayılır)
    D: MERKEZ  E: LÜLEBURGAZ  F: KAPAKLI   (işaret: x/X = kabul)
"""

import argparse
import re

from openpyxl import load_workbook


def digits(s: str) -> str:
    return re.sub(r"[^0-9]", "", s or "")


def truthy(cell: object) -> bool:
    return str(cell or "").strip().lower() == "x"


def parse_rows(path: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    kayitlar: list[dict] = []
    bolum_ctx: str | None = None
    grup_ctx: str | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        kod_raw, tanim, _aciklama, merkez, lule, kapakli = (list(row) + [None] * 6)[:6]
        if not (kod_raw and str(kod_raw).strip()):
            continue

        kod = str(kod_raw).strip()
        tanim = str(tanim).strip() if tanim else ""
        d = digits(kod)

        if len(d) <= 2:
            bolum_ctx, grup_ctx = tanim, None
            continue
        if len(d) == 4:
            grup_ctx = tanim
            continue
        if len(d) != 6:
            continue

        kayitlar.append({
            "kod": kod,
            "kod_temiz": d,
            "tanim": tanim,
            "tehlikeli": "*" in kod,
            "merkez": truthy(merkez),
            "luleburgaz": truthy(lule),
            "kapakli": truthy(kapakli),
            "bolum": bolum_ctx,
            "grup": grup_ctx,
        })
    return kayitlar


def load(kayitlar: list[dict]) -> None:
    from app.db import get_conn, init_db

    init_db()
    with get_conn() as conn:
        conn.execute("delete from atik_kodlari")
        conn.executemany(
            "insert into atik_kodlari "
            "(kod, kod_temiz, tanim, tehlikeli, merkez, luleburgaz, kapakli, bolum, grup) "
            "values (:kod, :kod_temiz, :tanim, :tehlikeli, "
            ":merkez, :luleburgaz, :kapakli, :bolum, :grup)",
            kayitlar,
        )
        conn.commit()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--dry-run", action="store_true", help="yükleme yapma, sadece say")
    args = parser.parse_args()

    kayitlar = parse_rows(args.path)
    tehlikeli = sum(k["tehlikeli"] for k in kayitlar)
    kabul = sum(k["merkez"] or k["luleburgaz"] or k["kapakli"] for k in kayitlar)
    print(f"Ayrıştırılan 6-haneli kod: {len(kayitlar)} | tehlikeli: {tehlikeli} | "
          f"en az bir tesiste kabul: {kabul}")

    if args.dry_run:
        return

    load(kayitlar)
    print(f"Yüklendi: {len(kayitlar)} kayıt → atik_kodlari")


if __name__ == "__main__":
    main()
